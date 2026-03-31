#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OSINT 自動化工作流程：從 URL/網域清單提取 Registrar 與 Creation Date。
使用 python-whois，無付費 API。支援緩存、日期正規化、錯誤標註。
"""

import argparse
import csv
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

# 可選：使用 tldextract 精確取得註冊網域；若未安裝則用簡易規則
try:
    import tldextract  # type: ignore[import-untyped]
    HAS_TLDEXTRACT = True
except ImportError:
    HAS_TLDEXTRACT = False

try:
    import whois  # type: ignore[import-untyped]
except ImportError:
    whois = None

# ---------------------------------------------------------------------------
# 常數
# ---------------------------------------------------------------------------
UNKNOWN = "Unknown/Private"
OUTPUT_CSV_COLUMNS = [
    "Domain",
    "Registrar",
    "Creation_Date",
    "Last_Checked_Timestamp",
    "Error_Reason",
]
# 常見 TLD（用於簡易網域提取）
COMMON_TLDS = frozenset(
    "com org net edu gov mil int info biz name co uk au de fr jp tw cn".split()
)
# 日期格式正則：多種 WHOIS 常見格式
DATE_PATTERNS = [
    (re.compile(r"(\d{4})-(\d{2})-(\d{2})"), lambda m: f"{m.group(1)}-{m.group(2)}-{m.group(3)}"),
    (re.compile(r"(\d{2})[-/.](\d{2})[-/.](\d{4})"), lambda m: f"{m.group(3)}-{m.group(2)}-{m.group(1)}"),
    (re.compile(r"(\d{4})[-/.](\d{2})[-/.](\d{2})"), lambda m: f"{m.group(1)}-{m.group(2)}-{m.group(3)}"),
    (re.compile(r"(\d{2})-([A-Za-z]{3})-(\d{4})"), lambda m: _parse_dd_mon_yyyy(m)),
    (re.compile(r"([A-Za-z]{3})\s+(\d{2}),?\s+(\d{4})"), lambda m: _parse_mon_dd_yyyy(m)),
]
MONTH_MAP = {
    "jan": "01", "feb": "02", "mar": "03", "apr": "04", "may": "05", "jun": "06",
    "jul": "07", "aug": "08", "sep": "09", "oct": "10", "nov": "11", "dec": "12",
}


# ---------------------------------------------------------------------------
# 網域提取
# ---------------------------------------------------------------------------
def extract_domain_from_url(url_or_domain: str) -> str:
    """
    從 URL 或已是網域的字串中提取「主網域」(eTLD+1)。
    例：https://www.google.com/search?q=google.com -> google.com
    """
    s = (url_or_domain or "").strip().lower()
    if not s:
        return ""

    # 若沒有 scheme，先當成 host 或網域
    if "://" not in s:
        return _get_registrable_domain(s)

    parsed = urlparse(s)
    host = (parsed.netloc or parsed.path or s).split("/")[0].split("?")[0]
    if not host:
        return ""
    return _get_registrable_domain(host)


def _get_registrable_domain(host: str) -> str:
    """從 host（如 www.google.com）取得註冊網域（如 google.com）。"""
    if HAS_TLDEXTRACT:
        ext = tldextract.extract(host)  # type: ignore[possibly-unbound]
        return ext.top_domain_under_public_suffix or host
    # 簡易規則：去掉常見子網域前綴，再取至少 domain.tld
    parts = [p for p in host.split(".") if p]
    if len(parts) <= 2:
        return host
    # 若最後一段為常見 TLD，取最後兩段
    if parts[-1] in COMMON_TLDS:
        return ".".join(parts[-2:])
    # 三級 TLD 如 co.uk
    if len(parts) >= 3 and parts[-1] in ("uk", "jp", "au") and parts[-2] in ("co", "com", "org", "ac", "gov"):
        return ".".join(parts[-3:])
    return ".".join(parts[-2:])


# ---------------------------------------------------------------------------
# 日期正規化
# ---------------------------------------------------------------------------
def _parse_dd_mon_yyyy(m) -> str:
    mon = MONTH_MAP.get(m.group(2).lower()[:3], "01")
    return f"{m.group(3)}-{mon}-{m.group(1)}"


def _parse_mon_dd_yyyy(m) -> str:
    mon = MONTH_MAP.get(m.group(1).lower()[:3], "01")
    return f"{m.group(3)}-{mon}-{m.group(2)}"


def normalize_creation_date(value) -> str:
    """
    將 WHOIS 回傳的 creation_date 統一為 YYYY-MM-DD。
    若無法解析或為隱私保護，回傳 UNKNOWN。
    """
    if value is None:
        return UNKNOWN
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (list, tuple)):
        for v in value:
            if isinstance(v, datetime):
                return v.strftime("%Y-%m-%d")
            s = normalize_creation_date(v)
            if s != UNKNOWN:
                return s
        return UNKNOWN
    s = str(value).strip()
    if not s or s.lower() in ("n/a", "unknown", "private"):
        return UNKNOWN
    for pat, fmt in DATE_PATTERNS:
        mo = pat.search(s)
        if mo:
            try:
                return fmt(mo)
            except (IndexError, KeyError):
                continue
    return UNKNOWN


def _extract_date_from_raw(raw_text: str) -> str:
    """從原始 WHOIS 文字用正則抓 creation date 字串並正規化。"""
    if not raw_text:
        return UNKNOWN
    # 常見欄位名
    for key in ("Creation Date", "Created", "created", "Registration Time", "Domain Registration Date"):
        pat = re.compile(rf"{re.escape(key)}\s*[:\s]+([^\s\n]+(?:\s+[^\s\n]+)*)", re.I)
        mo = pat.search(raw_text)
        if mo:
            return normalize_creation_date(mo.group(1).strip())
    return UNKNOWN


# ---------------------------------------------------------------------------
# WHOIS 查詢與解析
# ---------------------------------------------------------------------------
def query_whois(domain: str) -> tuple[str, str, str]:
    """
    對單一網域執行 WHOIS 查詢。
    回傳 (registrar, creation_date, error_reason)。
    """
    if not whois:
        return UNKNOWN, UNKNOWN, "Missing python-whois; install with: pip install python-whois"

    domain = (domain or "").strip().lower()
    if not domain:
        return UNKNOWN, UNKNOWN, "Empty domain"

    try:
        w = whois.whois(domain)
    except Exception as e:
        err = str(e).lower()
        if "connection refused" in err:
            return UNKNOWN, UNKNOWN, "Connection Refused"
        if "rate limit" in err or "limit exceeded" in err:
            return UNKNOWN, UNKNOWN, "Rate Limited"
        return UNKNOWN, UNKNOWN, f"Error: {e}"

    # Registrar
    reg = getattr(w, "registrar", None)
    if reg is None and hasattr(w, "registrar_name"):
        reg = getattr(w, "registrar_name", None)
    if reg is None and hasattr(w, "raw"):
        mo = re.search(r"Registrar:\s*(.+?)(?:\n|$)", (getattr(w, "raw", "") or ""), re.I)
        if mo:
            reg = mo.group(1).strip()
    registrar = (reg if reg else UNKNOWN)
    if isinstance(registrar, (list, tuple)):
        registrar = (registrar[0] if registrar else UNKNOWN)
    registrar = (registrar or UNKNOWN).strip() or UNKNOWN

    # Creation date：優先使用解析後的 creation_date，再試 raw
    creation = getattr(w, "creation_date", None)
    creation_str = normalize_creation_date(creation)
    if creation_str == UNKNOWN and getattr(w, "raw", None):
        creation_str = _extract_date_from_raw(getattr(w, "raw", ""))

    return registrar, creation_str, ""


# ---------------------------------------------------------------------------
# 緩存
# ---------------------------------------------------------------------------
class WhoisCache:
    """簡單緩存：key=domain, value=(registrar, creation_date, error_reason, timestamp)。"""

    def __init__(self, cache_file: str | Path | None = None):
        self._mem: dict[str, tuple[str, str, str, str]] = {}
        self._cache_file = Path(cache_file) if cache_file else None
        if self._cache_file and self._cache_file.exists():
            self._load()

    def _load(self) -> None:
        if self._cache_file is None:
            return
        try:
            with open(self._cache_file, "r", encoding="utf-8", newline="") as f:
                r = csv.DictReader(f)
                for row in r:
                    d = row.get("Domain", "").strip().lower()
                    if d:
                        self._mem[d] = (
                            row.get("Registrar", UNKNOWN),
                            row.get("Creation_Date", UNKNOWN),
                            row.get("Error_Reason", ""),
                            row.get("Last_Checked_Timestamp", ""),
                        )
        except Exception:
            pass

    def get(self, domain: str) -> tuple[str, str, str, str] | None:
        return self._mem.get(domain.strip().lower())

    def set(self, domain: str, registrar: str, creation_date: str, error_reason: str, ts: str) -> None:
        self._mem[domain.strip().lower()] = (registrar, creation_date, error_reason, ts)

    def save_to_csv(self, path: str | Path, rows: list[dict]) -> None:
        """向後相容：仍可寫 CSV 緩存檔。"""
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=OUTPUT_CSV_COLUMNS)
            w.writeheader()
            w.writerows(rows)


# ---------------------------------------------------------------------------
# Excel 輸出（每日一個 Sheet，累積不覆寫）
# ---------------------------------------------------------------------------
def save_to_excel(path: str | Path, rows: list[dict]) -> None:
    """
    將查詢結果寫入 Excel 檔案（.xlsx）。
    - 以今天日期（YYYY-MM-DD）為 Sheet 名稱。
    - 若檔案已存在，載入舊資料；若同日期的 Sheet 已存在，新資料附加在後面。
    - 不同日期會開新的 Sheet。
    """
    from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
    from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
    from openpyxl.styles import Font, Alignment, PatternFill  # type: ignore[import-untyped]

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    today_sheet = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d")

    # 載入或新建 Workbook
    if path.exists():
        wb = load_workbook(str(path))
    else:
        wb = Workbook()
        # 移除預設空白 Sheet
        default_sheet = wb.active
        if default_sheet is not None and default_sheet.title == "Sheet":
            wb.remove(default_sheet)

    # 取得或建立今天的 Sheet
    if today_sheet in wb.sheetnames:
        ws = wb[today_sheet]
    else:
        ws = wb.create_sheet(title=today_sheet)
        # 寫入標題列
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col_idx, col_name in enumerate(OUTPUT_CSV_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # 附加新資料（從最後一行之後開始）
    start_row = ws.max_row + 1
    for row_data in rows:
        for col_idx, col_name in enumerate(OUTPUT_CSV_COLUMNS, start=1):
            ws.cell(row=start_row, column=col_idx, value=row_data.get(col_name, ""))
        start_row += 1

    # 自動調整欄寬
    for col_idx in range(1, len(OUTPUT_CSV_COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    try:
        wb.save(str(path))
    except PermissionError:
        # 檔案可能正被 Excel 開啟，改存到帶時間戳的備用檔
        ts = datetime.now(tz=timezone.utc).strftime("%H%M%S")
        fallback = path.with_name(f"{path.stem}_{today_sheet}_{ts}{path.suffix}")
        print(f"[警告] 無法寫入 {path}（檔案可能被 Excel 開啟中），改存至 {fallback}", file=sys.stderr)
        wb.save(str(fallback))
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------
def run(
    inputs: list[str],
    output_file: str | Path = "whois_results.xlsx",
    cache_file: str | Path | None = "whois_cache.csv",
    use_cache: bool = True,
) -> list[dict]:
    """
    輸入 URL 或網域清單，查詢 WHOIS，寫入 Excel（每日一個 Sheet，累積不覆寫）。
    回傳所有結果列（list of dict）供進一步使用。
    """
    cache = WhoisCache(cache_file) if use_cache else WhoisCache(None)
    now_ts = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    results = []
    seen_domains: set[str] = set()

    for raw in inputs:
        domain = extract_domain_from_url(raw)
        if not domain:
            results.append({
                "Domain": raw[:200],
                "Registrar": UNKNOWN,
                "Creation_Date": UNKNOWN,
                "Last_Checked_Timestamp": now_ts,
                "Error_Reason": "Could not extract domain from input",
            })
            continue
        if domain in seen_domains:
            # 仍從緩存取一筆代表
            cached = cache.get(domain)
            if cached:
                results.append({
                    "Domain": domain,
                    "Registrar": cached[0],
                    "Creation_Date": cached[1],
                    "Last_Checked_Timestamp": cached[3] or now_ts,
                    "Error_Reason": cached[2],
                })
            continue
        seen_domains.add(domain)

        # 先查緩存
        if use_cache:
            cached = cache.get(domain)
            if cached:
                results.append({
                    "Domain": domain,
                    "Registrar": cached[0],
                    "Creation_Date": cached[1],
                    "Last_Checked_Timestamp": cached[3] or now_ts,
                    "Error_Reason": cached[2],
                })
                continue

        registrar, creation_date, error_reason = query_whois(domain)
        cache.set(domain, registrar, creation_date, error_reason, now_ts)
        results.append({
            "Domain": domain,
            "Registrar": registrar,
            "Creation_Date": creation_date,
            "Last_Checked_Timestamp": now_ts,
            "Error_Reason": error_reason,
        })

    # 寫入 Excel（累積模式）
    out_path = Path(output_file)
    save_to_excel(out_path, results)

    # 緩存仍用 CSV（僅供加速重複查詢）— 保存全部累積的緩存資料
    if cache_file and use_cache:
        cache._cache_file = Path(cache_file)
        all_cached_rows = [
            {
                "Domain": d,
                "Registrar": v[0],
                "Creation_Date": v[1],
                "Error_Reason": v[2],
                "Last_Checked_Timestamp": v[3],
            }
            for d, v in cache._mem.items()
        ]
        cache.save_to_csv(cache_file, all_cached_rows)

    return results


def main() -> None:
    parser = argparse.ArgumentParser(description="WHOIS 網域註冊商與註冊時間查詢（OSINT）")
    parser.add_argument("input", nargs="*", help="URL 或網域（可多個）")
    parser.add_argument("-o", "--output", default="whois_results.xlsx", help="輸出 Excel 路徑（.xlsx）")
    parser.add_argument("--cache-file", default="whois_cache.csv", help="緩存 CSV 路徑")
    parser.add_argument("--no-cache", action="store_true", help="停用緩存")
    parser.add_argument("-f", "--file", help="從檔案讀入 URL/網域清單（一行一筆）")
    args = parser.parse_args()

    inputs = list(args.input)
    if args.file:
        p = Path(args.file)
        if p.exists():
            inputs.extend(p.read_text(encoding="utf-8", errors="ignore").strip().splitlines())

    if not inputs:
        # 無輸入時使用範例 URL，方便在 Cursor 直接按「執行」
        inputs = ["https://www.google.com/search?q=google.com"]
        print("未提供輸入，使用範例 URL: https://www.google.com/search?q=google.com", file=sys.stderr)

    results = run(
        inputs,
        output_file=args.output,
        cache_file=args.cache_file,
        use_cache=not args.no_cache,
    )
    print(f"已寫入 {len(results)} 筆至 {args.output}")
    for r in results:
        err = f" [{r['Error_Reason']}]" if r.get("Error_Reason") else ""
        print(f"  {r['Domain']} -> Registrar: {r['Registrar']}, Creation: {r['Creation_Date']}{err}")


if __name__ == "__main__":
    main()
